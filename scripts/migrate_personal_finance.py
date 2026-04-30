#!/usr/bin/env python3
"""
Migrates PersonalFinance.xlsx to the new consolidated single-file format.

Usage:
    python scripts/migrate_personal_finance.py

Input:  data/PersonalFinance.xlsx
Output: data/PersonalFinance_migrated.xlsx

─── Sheet layout of the new file ─────────────────────────────────────────────

INPUT SHEETS  (where you type data each month):

  US Monthly    Date | Cash | Taxable Brokerage | Roth IRA | 401k | HSA |
                US Spend | Net Override | [Net USD*] | Notes

  UK Monthly    Date | Cash (£) | ETFs (£) | Pension (£) |
                UK Spend (£) | GBP/USD | Net Override (£) | [Net GBP*] | Notes

  Paychecks     Date | Gross | Pension Contrib | Net | GBP/USD | Notes

  US Asset Allocation  (overwrite monthly — current snapshot)
  UK Asset Allocation  (overwrite monthly — current snapshot)

  * Green columns are formula-computed. Do not edit them directly.

OUTPUT SHEETS  (formula-driven via Google Sheets ARRAYFORMULA — do not edit):

  US Networth, UK Networth, US Spend, UK Spend, Total Comp

  These are read by the dashboard. They auto-update when you add rows to the
  input sheets. Download as .xlsx to upload to the dashboard.

─── Monthly workflow ──────────────────────────────────────────────────────────

  1. Add 1 row to US Monthly  (balances + spend)
  2. Add 1 row to UK Monthly  (balances + spend)
  3. Add N rows to Paychecks  (one per payment received that month)
  4. Overwrite asset allocation sheets if you want holdings updated
  5. File → Download → Microsoft Excel (.xlsx) → upload to dashboard

─── Historical data ───────────────────────────────────────────────────────────

  Historical net worth totals are migrated into the 'Net Override' column
  (yellow background). For these rows the per-account breakdown columns
  (Cash, Taxable, etc.) are left as 0 — we can't reconstruct them.

  For new monthly entries: fill in the account columns and leave
  Net Override blank. The [Net USD] / [Net GBP] formula uses the account
  sum when Net Override is blank.
"""

import sys
from calendar import monthrange
from datetime import datetime, timedelta, date
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── constants ─────────────────────────────────────────────────────────────────

SRC = Path("data/PersonalFinance.xlsx")
DST = Path("data/PersonalFinance_migrated.xlsx")

EXCEL_EPOCH = datetime(1899, 12, 30)

# Header fill colours
FILL_INPUT   = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")  # blue
FILL_FORMULA = PatternFill(start_color="375623", end_color="375623", fill_type="solid")  # green
FILL_LEGACY  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow

FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_LEGACY_NOTE = Font(italic=True, color="7F7F7F", size=9)

DATE_FMT    = "YYYY-MM-DD"
DECIMAL_FMT = "#,##0.00"

# ── xlsx reader (no openpyxl dependency for reading — avoids formula eval) ────

def _shared_strings(z: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings: list[str] = []
    for si in ET.parse(z.open("xl/sharedStrings.xml")).getroot().findall("x:si", ns):
        t = si.find("x:t", ns)
        if t is not None:
            strings.append(t.text or "")
        else:
            strings.append("".join(
                r.find("x:t", ns).text or ""
                for r in si.findall("x:r", ns)
                if r.find("x:t", ns) is not None
            ))
    return strings


def _sheet_path(z: zipfile.ZipFile, sheet_name: str) -> str | None:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for s in ET.parse(z.open("xl/workbook.xml")).getroot().findall(".//x:sheet", ns):
        if s.get("name") == sheet_name:
            rid = s.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            for r in ET.parse(z.open("xl/_rels/workbook.xml.rels")).getroot():
                if r.get("Id") == rid:
                    target = r.get("Target")
                    return f"xl/{target}" if not target.startswith("/") else target[1:]
    return None


def read_sheet(z: zipfile.ZipFile, sheet_name: str, strings: list[str], skip_header: bool = True) -> list[list]:
    path = _sheet_path(z, sheet_name)
    if path is None:
        return []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rows: list[list] = []
    for i, row in enumerate(ET.parse(z.open(path)).getroot().findall(".//x:row", ns)):
        if skip_header and i == 0:
            continue
        cells = []
        for c in row.findall("x:c", ns):
            v = c.find("x:v", ns)
            if v is None:
                cells.append(None)
            elif c.get("t") == "s":
                cells.append(strings[int(v.text)])
            else:
                cells.append(v.text)
        if any(x is not None for x in cells):
            rows.append(cells)
    return rows

# ── value helpers ─────────────────────────────────────────────────────────────

def to_float(v) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def serial_to_date(serial) -> date | None:
    if serial is None:
        return None
    try:
        return (EXCEL_EPOCH + timedelta(days=int(float(serial)))).date()
    except (ValueError, TypeError):
        return None


def get(row: list, idx: int, default=None):
    return row[idx] if len(row) > idx else default


# ── month-key normalisation ───────────────────────────────────────────────────
#
# The old file stored networth on month-start (e.g. 2026-05-01) and spending
# on month-end (e.g. 2026-04-30).  Both refer to the same calendar month, so
# we normalise everything to a (year, month) key and emit one row per month.
#
# Rule: a networth date recorded on day 1–5 is treated as a late snapshot of
# the PREVIOUS month (the user checked balances just after month-end).
# Spending dates always belong to the month they fall in.

def _prev_month(year: int, month: int) -> tuple[int, int]:
    return (year - 1, 12) if month == 1 else (year, month - 1)


def _nw_month_key(dt: date) -> tuple[int, int]:
    if dt.day <= 5:
        return _prev_month(dt.year, dt.month)
    return (dt.year, dt.month)


def _sp_month_key(dt: date) -> tuple[int, int]:
    return (dt.year, dt.month)

# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _header_cell(ws, row: int, col: int, value: str, fill: PatternFill) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEADER
    c.fill = fill
    c.alignment = Alignment(horizontal="center", vertical="center")


def write_headers(ws, headers: list[tuple[str, PatternFill]]) -> None:
    ws.row_dimensions[1].height = 22
    for col_idx, (label, fill) in enumerate(headers, start=1):
        _header_cell(ws, 1, col_idx, label, fill)
    ws.freeze_panes = "A2"


def date_cell(ws, row: int, col: int, value: date | None) -> None:
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = DATE_FMT


def num_cell(ws, row: int, col: int, value: float | None, fmt: str = DECIMAL_FMT) -> None:
    c = ws.cell(row=row, column=col, value=value)
    if value is not None:
        c.number_format = fmt


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ── input sheet builders ──────────────────────────────────────────────────────

def build_us_monthly(wb, us_nw_by_serial: dict, us_sp_by_serial: dict) -> None:
    ws = wb.create_sheet("US Monthly")

    headers = [
        ("Date",               FILL_INPUT),
        ("Cash",               FILL_INPUT),
        ("Taxable Brokerage",  FILL_INPUT),
        ("Roth IRA",           FILL_INPUT),
        ("401k",               FILL_INPUT),
        ("HSA",                FILL_INPUT),
        ("US Spend",           FILL_INPUT),
        ("Net Override",       FILL_INPUT),
        ("Net USD",            FILL_FORMULA),   # col I — do not edit
        ("Notes",              FILL_INPUT),
    ]
    write_headers(ws, headers)

    # Normalise to (year, month) keys — one row per month.
    # If multiple NW entries fall in the same month (old irregular recording),
    # keep the one with the largest serial (most recent snapshot).
    nw_by_month: dict[tuple, tuple] = {}
    for serial, (net, conv) in us_nw_by_serial.items():
        dt = serial_to_date(serial)
        if dt is None:
            continue
        key = _nw_month_key(dt)
        if key not in nw_by_month or float(serial) > float(nw_by_month[key][2]):
            nw_by_month[key] = (net, conv, serial)

    sp_by_month: dict[tuple, float] = {}
    for serial, spend in us_sp_by_serial.items():
        dt = serial_to_date(serial)
        if dt is None:
            continue
        sp_by_month[_sp_month_key(dt)] = spend

    all_months = sorted(set(nw_by_month) | set(sp_by_month), reverse=True)

    for row_num, (year, month) in enumerate(all_months, start=2):
        dt = date(year, month, monthrange(year, month)[1])
        nw = nw_by_month.get((year, month))
        spend = sp_by_month.get((year, month), 0.0)
        net_override = nw[0] if nw else None

        date_cell(ws, row_num, 1, dt)
        num_cell(ws, row_num, 2, 0.0)   # Cash
        num_cell(ws, row_num, 3, 0.0)   # Taxable Brokerage
        num_cell(ws, row_num, 4, 0.0)   # Roth IRA
        num_cell(ws, row_num, 5, 0.0)   # 401k
        num_cell(ws, row_num, 6, 0.0)   # HSA
        num_cell(ws, row_num, 7, spend) # US Spend

        if net_override is not None:
            c = ws.cell(row=row_num, column=8, value=net_override)
            c.number_format = DECIMAL_FMT
            c.fill = FILL_LEGACY

        ws.cell(row=row_num, column=9,
                value=f"=IF(H{row_num}>0,H{row_num},B{row_num}+C{row_num}+D{row_num}+E{row_num}+F{row_num})")

    set_col_widths(ws, {
        "A": 14, "B": 12, "C": 20, "D": 12, "E": 10,
        "F": 10, "G": 12, "H": 16, "I": 14, "J": 30,
    })


def build_uk_monthly(wb, uk_nw_by_serial: dict, uk_sp_by_serial: dict) -> None:
    ws = wb.create_sheet("UK Monthly")

    headers = [
        ("Date",             FILL_INPUT),
        ("Cash (£)",         FILL_INPUT),
        ("ETFs (£)",         FILL_INPUT),
        ("Pension (£)",      FILL_INPUT),
        ("UK Spend (£)",     FILL_INPUT),
        ("GBP/USD",          FILL_INPUT),
        ("Net Override (£)", FILL_INPUT),
        ("Net GBP",          FILL_FORMULA),   # col H — do not edit
        ("Notes",            FILL_INPUT),
    ]
    write_headers(ws, headers)

    nw_by_month: dict[tuple, tuple] = {}
    for serial, (net, conv) in uk_nw_by_serial.items():
        dt = serial_to_date(serial)
        if dt is None:
            continue
        key = _nw_month_key(dt)
        if key not in nw_by_month or float(serial) > float(nw_by_month[key][2]):
            nw_by_month[key] = (net, conv, serial)

    sp_by_month: dict[tuple, tuple] = {}
    for serial, (spend, conv) in uk_sp_by_serial.items():
        dt = serial_to_date(serial)
        if dt is None:
            continue
        sp_by_month[_sp_month_key(dt)] = (spend, conv)

    all_months = sorted(set(nw_by_month) | set(sp_by_month), reverse=True)

    for row_num, (year, month) in enumerate(all_months, start=2):
        dt = date(year, month, monthrange(year, month)[1])
        nw = nw_by_month.get((year, month))
        sp = sp_by_month.get((year, month))

        net_override_gbp = nw[0] if nw else None
        gbp_usd = (nw[1] if nw else None) or (sp[1] if sp else 1.0)
        spend_gbp = sp[0] if sp else 0.0

        date_cell(ws, row_num, 1, dt)
        num_cell(ws, row_num, 2, 0.0)
        num_cell(ws, row_num, 3, 0.0)
        num_cell(ws, row_num, 4, 0.0)
        num_cell(ws, row_num, 5, spend_gbp)
        num_cell(ws, row_num, 6, gbp_usd, fmt="0.0000")

        if net_override_gbp is not None:
            c = ws.cell(row=row_num, column=7, value=net_override_gbp)
            c.number_format = DECIMAL_FMT
            c.fill = FILL_LEGACY

        ws.cell(row=row_num, column=8,
                value=f"=IF(G{row_num}>0,G{row_num},B{row_num}+C{row_num}+D{row_num})")

    set_col_widths(ws, {
        "A": 14, "B": 12, "C": 12, "D": 14,
        "E": 14, "F": 10, "G": 18, "H": 12, "I": 30,
    })


def build_paychecks(wb, raw_comp: list[list]) -> None:
    ws = wb.create_sheet("Paychecks")

    headers = [
        ("Date",                 FILL_INPUT),
        ("Gross",                FILL_INPUT),
        ("Pension Contrib",      FILL_INPUT),
        ("Approx Tax Reserves",  FILL_INPUT),
        ("Net",                  FILL_INPUT),
        ("GBP/USD",              FILL_INPUT),
        ("Notes",                FILL_INPUT),
    ]
    write_headers(ws, headers)

    for row_num, row in enumerate(raw_comp, start=2):
        if not row or get(row, 0) is None:
            continue
        dt          = serial_to_date(get(row, 0))
        gross       = to_float(get(row, 1))
        pension     = to_float(get(row, 2))
        tax_reserve = to_float(get(row, 3))
        net         = to_float(get(row, 4))
        gbp_usd     = to_float(get(row, 5)) or 1.0

        date_cell(ws, row_num, 1, dt)
        num_cell(ws, row_num, 2, gross)
        num_cell(ws, row_num, 3, pension)
        num_cell(ws, row_num, 4, tax_reserve)
        num_cell(ws, row_num, 5, net)
        num_cell(ws, row_num, 6, gbp_usd, fmt="0.0000")

    set_col_widths(ws, {
        "A": 14, "B": 16, "C": 18, "D": 22, "E": 16, "F": 10, "G": 30,
    })


def build_asset_sheet(wb, name: str, headers: list[str], raw: list[list]) -> None:
    ws = wb.create_sheet(name)
    for col_idx, h in enumerate(headers, start=1):
        _header_cell(ws, 1, col_idx, h, FILL_INPUT)
    ws.freeze_panes = "A2"

    for row_num, row in enumerate(raw, start=2):
        for col_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            f = to_float(val)
            ws.cell(row=row_num, column=col_idx, value=f if f is not None else val)

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 24

# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if not SRC.exists():
        print(f"Error: {SRC} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {SRC} ...")
    with zipfile.ZipFile(SRC) as z:
        ss = _shared_strings(z)
        raw_us_nw = read_sheet(z, "US Networth",         ss)
        raw_uk_nw = read_sheet(z, "UK Networth",         ss)
        raw_us_sp = read_sheet(z, "US Spend",            ss)
        raw_uk_sp = read_sheet(z, "UK Spend",            ss)
        raw_comp  = read_sheet(z, "Total Comp",          ss)
        raw_us_aa = read_sheet(z, "US Asset Allocation", ss)
        raw_uk_aa = read_sheet(z, "UK Asset Allocation", ss)

    # build lookup dicts keyed by Excel serial string (preserves original key type)
    us_nw_by_serial = {r[0]: (to_float(r[1]), to_float(get(r, 2)) or 1.0)
                       for r in raw_us_nw if r and r[0] is not None}
    uk_nw_by_serial = {r[0]: (to_float(r[1]), to_float(get(r, 2)) or 1.0)
                       for r in raw_uk_nw if r and r[0] is not None}
    us_sp_by_serial = {r[0]: to_float(r[1]) or 0.0
                       for r in raw_us_sp if r and r[0] is not None}
    uk_sp_by_serial = {r[0]: (to_float(r[1]) or 0.0, to_float(get(r, 2)) or 1.0)
                       for r in raw_uk_sp if r and r[0] is not None}

    print(f"  US Networth: {len(us_nw_by_serial)} rows")
    print(f"  UK Networth: {len(uk_nw_by_serial)} rows")
    print(f"  US Spend:    {len(us_sp_by_serial)} rows")
    print(f"  UK Spend:    {len(uk_sp_by_serial)} rows")
    print(f"  Paychecks:   {len(raw_comp)} rows")
    print(f"  US Alloc:    {len(raw_us_aa)} rows")
    print(f"  UK Alloc:    {len(raw_uk_aa)} rows")

    print(f"\nBuilding {DST} ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── input sheets ──
    build_us_monthly(wb, us_nw_by_serial, us_sp_by_serial)
    build_uk_monthly(wb, uk_nw_by_serial, uk_sp_by_serial)
    build_paychecks(wb, raw_comp)
    build_asset_sheet(wb, "US Asset Allocation",
                      ["Account Type", "Asset", "Value", "Conversion"], raw_us_aa)
    build_asset_sheet(wb, "UK Asset Allocation",
                      ["Account Type", "Asset", "Value", "Conversion"], raw_uk_aa)

    wb.save(DST)

    print(f"\nDone. Sheets in {DST}:")
    for name in wb.sheetnames:
        print(f"  {name}")

    print("""
Next steps:
  1. Upload data/PersonalFinance_migrated.xlsx to Google Drive
  2. Open it and choose  File → Save as Google Sheets
  3. Confirm the output sheet formulas evaluate (they need Google Sheets to run)
  4. For new monthly entries: add rows to US Monthly / UK Monthly / Paychecks
  5. To use with the dashboard: File → Download → Microsoft Excel (.xlsx)
""")


if __name__ == "__main__":
    main()
